from LSvehicleOneDimLookup_2 import vehicleOneDimLookup_2 as vehicle_ODL_2
from LSvehicleTwoDimLookup_2 import vehicleTwoDimLookup_2 as vehicle_TDL_2
from Create_TrackMap_2D import Create_TrackMap_2D
from dask.array.core import asarray
from eventPoints import Scoring
from eventPoints import Event
from track import Track
from Plots import Plots
import pandas as pandas
import lapsim as lts
import pylab as pl
import numpy as np
import platform
import openpyxl
import math
import os
from openpyxl import Workbook

class ParamStudy:
    def __init__(self, EventName, Year, Parameter, Stepsize, MinValue, MaxValue,car_type,track,simulation_name):
        
        self.EventName = EventName
        self.Year = Year
        self.Parameter = Parameter
        self.Stepsize = Stepsize
        self.StartVehicle = 'FSG15e2D'
        self.car_type = car_type
        self.track = os.listdir(os.getcwd()+"/tracks/"+str(track))
        self.track.append ("Skidpad")
        self.track.append("Acceleration")    #['Endurance_FSA15_Curvature', 'AutoX_FSA15_Curvature', 'Skidpad', 'Acceleration', 'Efficiency_FSA15_Curvature']
        self.track_name = track
        print(self.track)
        self.MinValue = MinValue
        self.MaxValue = MaxValue
        self.StartValue = 0
        self.Counter = 1
        self.Pmax = 80000
        self.simulation_name =simulation_name


    def SimulateParamStudy(self):
        #Load Vehicle Parameters from file
        cwd = os.getcwd()
        if('Windows' in platform.platform()):
            file = cwd + "/car_modells/" + self.car_type
        else:
            cwd = cwd.replace('\\', '/')
            file = cwd + '/CSV/VehicleParameters.xlsx'
            
        wb = openpyxl.load_workbook(file)
        car_data = wb.active
        self.car_data_loaded = []

        #Generate file to save at
        self.result_file = Workbook()
        self.result_file.title = "Results"
        self.result_file_sheet = self.result_file.active

        #Init File Header
        first_line = ["Parameter: "+str(self.Parameter),"AutoX Score","AutoX Time","Efficiency Score","Efficiency Time","Endurance Score","Endurance Time","Skidpad Score","Skidpad Time","Acceleration Score","Acceleration Time","Overall"]
        self.result_file_sheet.append(first_line)
        self.result_file_name = os.getcwd() + "/Results/"+ str(self.simulation_name)+"_simulation_Results.xlsx"
        self.result_file.save(filename = self.result_file_name)

        for a in range(1,16):

            self.car_data_loaded.append(car_data.cell(row=a,column=1).value)
        print(self.car_data_loaded)
        if ('FSG15e2D' in self.StartVehicle):
            FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
            self.StartVehicle = FSG15e2D    #initialize two dim vehicle
            self.StartValue = self.MinValue

    
            while (self.StartValue <= self.MaxValue):   #loop for increase the parameter value
                if(self.Parameter == 'C_F'):
                    self.car_data_loaded[0] = self.StartValue
                    #initialize vehicle with new value
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded,'FSG15e2D')
                elif(self.Parameter == 'C_R'):
                    self.car_data_loaded[1] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'm'):
                    self.car_data_loaded[2] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'CoG_X'):
                    self.car_data_loaded[3] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'mu'):
                    self.car_data_loaded[4] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'alpha'):
                    self.car_data_loaded[5] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'CoP_X'):
                    self.car_data_loaded[6] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'C_la'):
                    self.car_data_loaded[7] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'rho'):
                    self.car_data_loaded[8] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'gearRatio'):
                    self.car_data_loaded[10] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'tireRadius'):
                    self.car_data_loaded[11] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'fr'):
                    self.car_data_loaded[12] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                elif(self.Parameter == 'Lift2Drag'):
                    self.car_data_loaded[13] = self.StartValue
                    FSG15e2D = vehicle_TDL_2(self.car_data_loaded, 'FSG15e2D')
                self.StartVehicle = FSG15e2D    #set start vehicle
                revision = round((int(self.MaxValue)-int(self.MinValue))/int(self.Stepsize+1),0)  #number of calculations to do

                #Write newParameter value
                print("Start_value")
                print(self.StartValue,self.Counter)
                self.result_file_sheet.cell(row=self.Counter+1,column=1).value =self.StartValue
                self.result_file.save(self.result_file_name)
                SimEvent = Event(self.EventName, self.Year, self.StartValue, self.Counter, revision, [self.StartVehicle], self.track,self.track_name,self.simulation_name,self.result_file_sheet,self.result_file_name,self.Counter+1)
                SimEvent.SimulateEvent()    #call function to simulate the event
                self.result_file.save(self.result_file_name)
                self.Counter += 1   #counter for result file (look module event points)
                self.StartValue += self.Stepsize    #increase the start value with stepsize
          

    
        #ResultPlots = Plots('Results.txt', self.MinValue, self.MaxValue, self.StartVehicle, self.x_label)
        #ResultPlots.ShowPlots()
